from typing import Any, Dict, Optional, List, Tuple
from datetime import datetime, timedelta
import logging
import io

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from firebase_admin import firestore
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.security import get_current_active_user
from app.config import COLLECTIONS
from app.database import get_db
from app.models import TokenData
from app.security import require_roles
from app.utils.responses import ok


router = APIRouter(prefix="/portal/pharmacy", tags=["Pharmacy (Role-Protected)"])

# Public pharmacy endpoints (mounted separately under /api/v1/pharmacy)
public_router = APIRouter(prefix="/pharmacy", tags=["Pharmacy"])


logger = logging.getLogger(__name__)


class AddMedicineRequest(BaseModel):
    product_id: int = Field(..., ge=0)
    batch_no: str
    name: str
    generic_name: str
    type: str
    distributor: str
    purchase_price: float = Field(..., gt=0)
    selling_price: float = Field(..., gt=0)
    stock_unit: str
    quantity: int = Field(..., ge=0)
    expiration_date: str
    category: str
    sub_category: str


class DispenseMedicineItem(BaseModel):
    product_id: int = Field(..., ge=0)
    quantity: int = Field(..., ge=1)


class DispenseMedicineRequest(BaseModel):
    patient_id: str
    doctor_id: str
    medicines: List[DispenseMedicineItem]


def _normalize_str(v: Optional[str]) -> str:
    return (v or "").strip()


def _normalize_date_str(v: Optional[str]) -> Optional[str]:
    s = _normalize_str(v)
    if not s:
        return None

    # Accept UI format DD/MM/YYYY
    try:
        if "/" in s and len(s.split("/")) == 3:
            dd, mm, yyyy = [p.strip() for p in s.split("/")]
            dt = datetime(int(yyyy), int(mm), int(dd))
            return dt.date().isoformat()
    except Exception:
        pass

    # Accept already-ISO date (YYYY-MM-DD) or full ISO datetime
    try:
        dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
        return dt.date().isoformat()
    except Exception:
        pass

    # Accept common format like "Dec 10, 2025"
    try:
        dt = datetime.strptime(s, "%b %d, %Y")
        return dt.date().isoformat()
    except Exception:
        pass

    # Accept common format like "10 Dec, 2025"
    try:
        dt = datetime.strptime(s, "%d %b, %Y")
        return dt.date().isoformat()
    except Exception:
        pass

    raise HTTPException(
        status_code=status.HTTP_400_BAD_REQUEST,
        detail="Date format is invalid. Please use DD/MM/YYYY",
    )


def _coerce_int(v: Any, *, field: str) -> int:
    try:
        return int(v)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{field} must be a number")


def _coerce_float(v: Any, *, field: str) -> float:
    try:
        return float(v)
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"{field} must be a number")


def _compute_status(item: Dict[str, Any]) -> str:
    existing = str(item.get("status") or "").strip().lower()
    if existing in {"active", "expired", "inactive"}:
        return existing
    expiry = item.get("expiry_date") or item.get("expiry")
    expiry_s = _normalize_str(str(expiry or ""))
    if not expiry_s:
        return "active" if bool(item.get("is_active", True)) else "inactive"
    try:
        exp_iso = _normalize_date_str(expiry_s)
        if not exp_iso:
            return "active" if bool(item.get("is_active", True)) else "inactive"
        exp_dt = datetime.fromisoformat(exp_iso)
        return "expired" if exp_dt.date() < datetime.utcnow().date() else "active"
    except Exception:
        return "active" if bool(item.get("is_active", True)) else "inactive"


def _to_datetime(value: Any) -> Optional[datetime]:
    if value is None:
        return None
    try:
        to_dt = getattr(value, "to_datetime", None)
        if callable(to_dt):
            return to_dt()
    except Exception:
        pass
    if isinstance(value, datetime):
        return value
    try:
        s = str(value).strip()
        if not s:
            return None
        # Accept date-only
        if len(s) == 10 and s[4] == "-" and s[7] == "-":
            return datetime.fromisoformat(s)
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


def _snap_exists(snap: Any) -> bool:
    try:
        ex = getattr(snap, "exists", None)
        if callable(ex):
            return bool(ex())
        return bool(ex)
    except Exception:
        return False


def _prefix_query(ref: Any, field: str, q: str, limit: int = 20) -> List[Dict[str, Any]]:
    """Firestore prefix search: field >= q and field <= q+\uf8ff.

    Uses order_by+start_at/end_at when available. Falls back to in-memory filtering.
    """
    qn = _normalize_str(q)
    if not qn:
        return []

    try:
        if hasattr(ref, "order_by") and hasattr(ref, "start_at") and hasattr(ref, "end_at"):
            ordered = ref.order_by(field)
            qr = ordered.start_at([qn]).end_at([qn + "\uf8ff"]).limit(int(limit))
            return [d.to_dict() for d in qr.stream()]
    except Exception:
        pass

    # Fallback: fetch a limited set and filter contains (case-insensitive)
    docs = [d.to_dict() for d in ref.limit(1000).stream()]
    ql = qn.lower()
    out: List[Dict[str, Any]] = []
    for it in docs:
        v = str(it.get(field) or "")
        if ql in v.lower():
            out.append(it)
        if len(out) >= limit:
            break
    return out


@public_router.get("/search-medicine")
async def search_medicine(
    q: str = Query(..., description="Search by medicine name or generic name"),
) -> Dict[str, Any]:
    db = get_db()
    medicines_ref = db.collection(COLLECTIONS.get("PHARMACY_MEDICINES") or "pharmacy_medicines")

    qn = _normalize_str(q)
    if not qn:
        return {"results": []}

    # Query name and generic_name separately, then merge unique by product_id
    by_name = _prefix_query(medicines_ref, "name", qn, limit=20)
    by_generic = _prefix_query(medicines_ref, "generic_name", qn, limit=20)

    merged: List[Dict[str, Any]] = []
    seen: set = set()
    for it in (by_name + by_generic):
        try:
            pid = int(it.get("product_id")) if it.get("product_id") is not None else None
        except Exception:
            pid = None
        key = pid if pid is not None else str(it.get("id") or it.get("doc_id") or "")
        if not key or key in seen:
            continue
        seen.add(key)
        merged.append(it)
        if len(merged) >= 20:
            break

    results: List[Dict[str, Any]] = []
    for it in merged[:20]:
        try:
            qty = int(it.get("quantity") or 0)
        except Exception:
            qty = 0
        try:
            selling = float(it.get("selling_price") or 0.0)
        except Exception:
            selling = 0.0
        exp = _to_datetime(it.get("expiration_date"))

        results.append(
            {
                "product_id": it.get("product_id"),
                "name": it.get("name"),
                "generic_name": it.get("generic_name"),
                "selling_price": selling,
                "quantity": qty,
                "expiration_date": exp.isoformat() + "Z" if exp else None,
                "low_stock": bool(qty < 5),
            }
        )

    return {"results": results}


@public_router.post("/dispense-medicine", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def dispense_medicine(
    payload: DispenseMedicineRequest,
    current: TokenData = Depends(get_current_active_user),
) -> Dict[str, Any]:
    """Deduct stock from pharmacy_medicines and record each dispensed line item into pharmacy_sales."""
    if not payload.medicines:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="medicines is required")

    db = get_db()
    meds_ref = db.collection(COLLECTIONS.get("PHARMACY_MEDICINES") or "pharmacy_medicines")
    sales_ref = db.collection(COLLECTIONS.get("PHARMACY_SALES") or "pharmacy_sales")

    # Normalize duplicates in request (same product_id)
    requested: Dict[int, int] = {}
    for it in payload.medicines:
        requested[int(it.product_id)] = requested.get(int(it.product_id), 0) + int(it.quantity)

    now = datetime.utcnow()
    user_id = getattr(current, "user_id", None)

    def _run_non_transactional() -> None:
        # Pre-check all items first to avoid partial updates
        loaded: Dict[int, Dict[str, Any]] = {}
        for pid, req_qty in requested.items():
            snap = meds_ref.document(str(pid)).get()
            if not _snap_exists(snap):
                raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Medicine not found")
            data = snap.to_dict() or {}
            loaded[pid] = data
            try:
                available = int(data.get("quantity") or 0)
            except Exception:
                available = 0
            if req_qty > available:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Insufficient stock")

        # Apply updates + sales
        for pid, req_qty in requested.items():
            data = loaded.get(pid) or {}
            try:
                available = int(data.get("quantity") or 0)
            except Exception:
                available = 0
            new_qty = max(0, available - int(req_qty))
            meds_ref.document(str(pid)).update({"quantity": new_qty})

            try:
                unit_price = float(data.get("selling_price") or 0.0)
            except Exception:
                unit_price = 0.0
            sale_doc = {
                "patient_id": payload.patient_id,
                "doctor_id": payload.doctor_id,
                "medicine_id": pid,
                "medicine_name": data.get("name"),
                "quantity": int(req_qty),
                "quantity_sold": int(req_qty),
                "unit_price": unit_price,
                "total_price": float(int(req_qty) * unit_price),
                "total_amount": float(int(req_qty) * unit_price),
                "payment_status": "paid",
                "sold_at": now,
                "dispensed_at": now,
                "created_at": now,
                "performed_by": user_id,
            }
            sref = sales_ref.document()
            sale_doc["id"] = sref.id
            sref.set(sale_doc)

    # Use Firestore transaction if available (atomic stock check + deduction)
    try:
        tx_factory = getattr(db, "transaction", None)
        if callable(tx_factory):
            transaction = db.transaction()

            @firestore.transactional
            def _txn_op(transaction):
                # Load and validate
                loaded: Dict[int, Dict[str, Any]] = {}
                for pid, req_qty in requested.items():
                    doc_ref = meds_ref.document(str(pid))
                    snap = doc_ref.get(transaction=transaction)
                    if not _snap_exists(snap):
                        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Medicine not found")
                    data = snap.to_dict() or {}
                    loaded[pid] = data
                    try:
                        available = int(data.get("quantity") or 0)
                    except Exception:
                        available = 0
                    if req_qty > available:
                        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Insufficient stock")

                # Update stock and create sales
                for pid, req_qty in requested.items():
                    data = loaded.get(pid) or {}
                    try:
                        available = int(data.get("quantity") or 0)
                    except Exception:
                        available = 0
                    new_qty = max(0, available - int(req_qty))
                    doc_ref = meds_ref.document(str(pid))
                    transaction.update(doc_ref, {"quantity": new_qty})

                    try:
                        unit_price = float(data.get("selling_price") or 0.0)
                    except Exception:
                        unit_price = 0.0
                    sale_doc = {
                        "patient_id": payload.patient_id,
                        "doctor_id": payload.doctor_id,
                        "medicine_id": pid,
                        "medicine_name": data.get("name"),
                        "quantity": int(req_qty),
                        "quantity_sold": int(req_qty),
                        "unit_price": unit_price,
                        "total_price": float(int(req_qty) * unit_price),
                        "total_amount": float(int(req_qty) * unit_price),
                        "payment_status": "paid",
                        "sold_at": now,
                        "dispensed_at": now,
                        "created_at": now,
                        "performed_by": user_id,
                    }
                    sref = sales_ref.document()
                    sale_doc["id"] = sref.id
                    transaction.set(sref, sale_doc)

            _txn_op(transaction)
        else:
            _run_non_transactional()

        logger.info(
            "Medicines dispensed",
            extra={"patient_id": payload.patient_id, "doctor_id": payload.doctor_id, "lines": len(requested), "user_id": user_id},
        )
        return {"success": True, "message": "Medicines dispensed successfully"}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Failed to dispense medicines")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to dispense medicines")


@public_router.post("/add-medicine", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def add_medicine(
    payload: AddMedicineRequest,
    current: TokenData = Depends(get_current_active_user),
) -> Dict[str, Any]:
    db = get_db()
    try:
        exp_iso = _normalize_date_str(payload.expiration_date)
        exp_dt = _to_datetime(exp_iso) if exp_iso else None
        if exp_dt is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid expiration_date")

        ref = db.collection("pharmacy_medicines").document(str(payload.product_id))
        snap = ref.get()
        if _snap_exists(snap):
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Medicine already exists")

        doc = {
            "product_id": int(payload.product_id),
            "batch_no": payload.batch_no,
            "name": payload.name,
            "generic_name": payload.generic_name,
            "type": payload.type,
            "distributor": payload.distributor,
            "purchase_price": float(payload.purchase_price),
            "selling_price": float(payload.selling_price),
            "stock_unit": payload.stock_unit,
            "quantity": int(payload.quantity),
            "expiration_date": exp_dt,
            "category": payload.category,
            "sub_category": payload.sub_category,
            "created_at": datetime.utcnow(),
        }

        ref.set(doc)
        logger.info("Medicine added", extra={"product_id": payload.product_id, "user_id": getattr(current, "user_id", None)})
        return {"success": True, "message": "Medicine added successfully"}
    except HTTPException:
        raise
    except Exception as e:
        logger.exception("Failed to add medicine")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail=str(e))


@public_router.get("/low-stock", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def low_stock_alerts(
    limit: int = Query(50, ge=1, le=200, description="Max number of low stock medicines to return"),
) -> Dict[str, Any]:
    db = get_db()
    medicines_ref = db.collection(COLLECTIONS.get("PHARMACY_MEDICINES") or "pharmacy_medicines")

    docs: List[Dict[str, Any]] = []
    try:
        q = medicines_ref.where("quantity", "<", 10)
        # Prefer consistent ordering when available
        if hasattr(q, "order_by"):
            try:
                q = q.order_by("quantity")
            except Exception:
                pass
        docs = [d.to_dict() for d in q.limit(int(limit)).stream()]
    except Exception:
        # Mock/in-memory fallback
        all_docs = [d.to_dict() for d in medicines_ref.limit(5000).stream()]
        filtered: List[Dict[str, Any]] = []
        for it in all_docs:
            try:
                qty = int(it.get("quantity") or 0)
            except Exception:
                qty = 0
            if qty < 10:
                filtered.append(it)
        filtered.sort(key=lambda x: int(x.get("quantity") or 0))
        docs = filtered[: int(limit)]

    results: List[Dict[str, Any]] = []
    for it in docs[: int(limit)]:
        try:
            qty = int(it.get("quantity") or 0)
        except Exception:
            qty = 0
        try:
            selling = float(it.get("selling_price") or 0.0)
        except Exception:
            selling = 0.0
        exp = _to_datetime(it.get("expiration_date"))

        results.append(
            {
                "product_id": it.get("product_id"),
                "name": it.get("name"),
                "generic_name": it.get("generic_name"),
                "selling_price": selling,
                "quantity": qty,
                "expiration_date": exp.isoformat() + "Z" if exp else None,
                "low_stock": True,
            }
        )

    return {"results": results}


def _period_range(period: str, start_date: Optional[str], end_date: Optional[str]) -> Tuple[Optional[datetime], Optional[datetime]]:
    p = _normalize_str(period).lower()
    now = datetime.utcnow()
    if p in {"today", "this_day", "day", "daily"}:
        start = datetime(now.year, now.month, now.day)
        return start, now
    if p in {"this_week", "week", "weekly"}:
        start = now - timedelta(days=7)
        return start, now
    if p in {"this_month", "month", "monthly"}:
        start = datetime(now.year, now.month, 1)
        return start, now
    if p in {"custom", "custom_range", "range"}:
        if not start_date or not end_date:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="start_date and end_date are required for custom range")
        s_iso = _normalize_date_str(start_date)
        e_iso = _normalize_date_str(end_date)
        if not s_iso or not e_iso:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="start_date and end_date are required for custom range")
        s_dt = datetime.fromisoformat(s_iso)
        e_dt = datetime.fromisoformat(e_iso) + timedelta(days=1)
        return s_dt, e_dt
    # all time
    return None, None


def _end_inclusive(end_exclusive: Optional[datetime]) -> Optional[datetime]:
    if end_exclusive is None:
        return None
    # Firestore supports <= but not <, so use an inclusive bound just before end_exclusive.
    return end_exclusive - timedelta(microseconds=1)


@router.get("/items", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def list_items(
    current: TokenData = Depends(get_current_active_user),
    hospital_id: Optional[str] = Query(None),
    q: Optional[str] = Query(None, description="Search by name or SKU"),
    sort: Optional[str] = Query(None, description="name_asc|name_desc|quantity_asc|quantity_desc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
) -> Any:
    db = get_db()
    items_ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"])

    query_ref = items_ref
    if hospital_id:
        query_ref = query_ref.where("hospital_id", "==", hospital_id)

    docs = [d.to_dict() for d in query_ref.stream()]
    docs = [d for d in docs if not _is_deleted(d)]

    q_norm = _normalize_str(q).lower()
    if q_norm:
        filtered = []
        for it in docs:
            name = str(it.get("name") or it.get("medicine_name") or "").lower()
            serial = str(it.get("serial_number") or it.get("sku") or "").lower()
            if q_norm in name or q_norm in serial:
                filtered.append(it)
        docs = filtered

    for it in docs:
        it["medicine_name"] = it.get("medicine_name") or it.get("name")
        it["serial_number"] = it.get("serial_number") or it.get("sku")
        it["batch_number"] = it.get("batch_number")
        it["manufacture_date"] = it.get("manufacture_date")
        it["expiry_date"] = it.get("expiry_date")
        it["quantity"] = int(it.get("quantity") if it.get("quantity") is not None else (it.get("stock") or 0))
        it["unit_price"] = it.get("unit_price") if it.get("unit_price") is not None else it.get("price")
        it["supplier_name"] = it.get("supplier_name")
        it["status"] = _compute_status(it)

    s = _normalize_str(sort).lower()
    if s == "name_asc":
        docs.sort(key=lambda x: str(x.get("medicine_name") or "").lower())
    elif s == "name_desc":
        docs.sort(key=lambda x: str(x.get("medicine_name") or "").lower(), reverse=True)
    elif s == "quantity_asc":
        docs.sort(key=lambda x: int(x.get("quantity") or 0))
    elif s == "quantity_desc":
        docs.sort(key=lambda x: int(x.get("quantity") or 0), reverse=True)

    total = len(docs)
    start = (page - 1) * page_size
    end = start + page_size
    items = docs[start:end]

    return ok(
        data=items,
        meta={"page": page, "page_size": page_size, "total": total},
    )


@router.post("/items", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def create_item(
    payload: Dict[str, Any],
    current: TokenData = Depends(get_current_active_user),
) -> Any:
    name = _normalize_str(str(payload.get("medicine_name") or payload.get("name") or ""))
    serial = _normalize_str(str(payload.get("serial_number") or payload.get("sku") or ""))
    batch_number = _normalize_str(str(payload.get("batch_number") or ""))
    hospital_id = _normalize_str(str(payload.get("hospital_id") or ""))

    if not name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Medicine name is required")
    if not serial:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Serial number is required")
    if not batch_number:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Batch number is required")
    if not hospital_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="hospital_id is required")

    quantity = _coerce_int(payload.get("quantity", payload.get("stock", 0)) or 0, field="quantity")
    if quantity < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Quantity cannot be negative")

    unit_price = _coerce_float(payload.get("unit_price", payload.get("price", 0)) or 0, field="unit_price")
    if unit_price < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Unit price cannot be negative")

    manufacture_date = _normalize_date_str(payload.get("manufacture_date"))
    expiry_date = _normalize_date_str(payload.get("expiry_date"))
    supplier_name = _normalize_str(str(payload.get("supplier_name") or ""))
    if not manufacture_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Manufacture date is required")
    if not expiry_date:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Expiry date is required")
    if not supplier_name:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Supplier name is required")

    db = get_db()
    items_ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"])

    # Best-effort uniqueness check (Firestore composite indexes may not exist; keep it simple)
    existing = [d.to_dict() for d in items_ref.where("hospital_id", "==", hospital_id).stream()]
    for it in existing:
        existing_serial = str(it.get("serial_number") or it.get("sku") or "").strip().lower()
        if existing_serial == serial.lower():
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Serial number already exists for this hospital")

    item_ref = items_ref.document()
    now = datetime.utcnow()
    item_data = {
        "id": item_ref.id,
        "hospital_id": hospital_id,
        "medicine_name": name,
        "name": name,
        "serial_number": serial,
        "sku": serial,
        "batch_number": batch_number,
        "manufacture_date": manufacture_date,
        "expiry_date": expiry_date,
        "supplier_name": supplier_name,
        "quantity": quantity,
        "stock": quantity,
        "unit_price": unit_price,
        "price": unit_price,
        "min_stock": _coerce_int(payload.get("min_stock") or 10, field="min_stock"),
        "status": _normalize_str(str(payload.get("status") or "")) or None,
        "is_active": bool(payload.get("is_active", True)),
        "created_by": current.user_id,
        "created_at": now,
        "updated_at": now,
    }

    item_ref.set(item_data)

    item_data["status"] = _compute_status(item_data)
    return ok(data=item_data, message="Medicine added to inventory")


@router.get("/items/{item_id}", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def get_item(
    item_id: str,
    current: TokenData = Depends(get_current_active_user),
) -> Any:
    db = get_db()
    doc = db.collection(COLLECTIONS["PHARMACY_ITEMS"]).document(item_id).get()
    if not getattr(doc, "exists", False):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")
    data = doc.to_dict() or {}
    data["medicine_name"] = data.get("medicine_name") or data.get("name")
    data["serial_number"] = data.get("serial_number") or data.get("sku")
    data["quantity"] = int(data.get("quantity") if data.get("quantity") is not None else (data.get("stock") or 0))
    data["unit_price"] = data.get("unit_price") if data.get("unit_price") is not None else data.get("price")
    data["status"] = _compute_status(data)
    return ok(data=data)


@router.put("/items/{item_id}", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def update_item(
    item_id: str,
    payload: Dict[str, Any],
    current: TokenData = Depends(get_current_active_user),
) -> Any:
    db = get_db()
    ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"]).document(item_id)
    snap = ref.get()
    if not getattr(snap, "exists", False):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")

    updates: Dict[str, Any] = {}

    if "medicine_name" in payload or "name" in payload:
        name = _normalize_str(str(payload.get("medicine_name") or payload.get("name") or ""))
        if not name:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Medicine name cannot be empty")
        updates["medicine_name"] = name
        updates["name"] = name

    if "serial_number" in payload or "sku" in payload:
        serial = _normalize_str(str(payload.get("serial_number") or payload.get("sku") or ""))
        if not serial:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Serial number cannot be empty")
        updates["serial_number"] = serial
        updates["sku"] = serial

    if "batch_number" in payload:
        bn = _normalize_str(str(payload.get("batch_number") or ""))
        if not bn:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Batch number cannot be empty")
        updates["batch_number"] = bn

    if "manufacture_date" in payload:
        updates["manufacture_date"] = _normalize_date_str(payload.get("manufacture_date"))

    if "expiry_date" in payload:
        updates["expiry_date"] = _normalize_date_str(payload.get("expiry_date"))

    if "supplier_name" in payload:
        sname = _normalize_str(str(payload.get("supplier_name") or ""))
        if not sname:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Supplier name cannot be empty")
        updates["supplier_name"] = sname

    if "unit_price" in payload or "price" in payload:
        updates["unit_price"] = _coerce_float(payload.get("unit_price", payload.get("price")), field="unit_price")
        updates["price"] = updates["unit_price"]

    if "status" in payload:
        updates["status"] = _normalize_str(str(payload.get("status") or "")) or None

    for key in ["is_active", "min_stock"]:
        if key in payload:
            updates[key] = payload.get(key)

    if "quantity" in payload or "stock" in payload:
        new_qty = _coerce_int(payload.get("quantity", payload.get("stock")), field="quantity")
        if new_qty < 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Quantity cannot be negative")
        current_stock = int((snap.to_dict() or {}).get("quantity") if (snap.to_dict() or {}).get("quantity") is not None else ((snap.to_dict() or {}).get("stock") or 0))
        delta = new_qty - current_stock
        updates["quantity"] = new_qty
        updates["stock"] = new_qty
        log_ref = db.collection(COLLECTIONS["PHARMACY_STOCK_LOGS"]).document()
        log_ref.set({
            "id": log_ref.id,
            "item_id": item_id,
            "hospital_id": (snap.to_dict() or {}).get("hospital_id"),
            "delta": delta,
            "from_stock": current_stock,
            "to_stock": new_qty,
            "reason": _normalize_str(str(payload.get("reason") or "manual_update")),
            "performed_by": current.user_id,
            "created_at": datetime.utcnow(),
        })

    updates["updated_at"] = datetime.utcnow()

    if not updates:
        return ok(data=snap.to_dict(), message="No changes to apply")

    ref.update(updates)
    new_data = snap.to_dict() or {}
    new_data.update(updates)

    new_data["medicine_name"] = new_data.get("medicine_name") or new_data.get("name")
    new_data["serial_number"] = new_data.get("serial_number") or new_data.get("sku")
    new_data["quantity"] = int(new_data.get("quantity") if new_data.get("quantity") is not None else (new_data.get("stock") or 0))
    new_data["unit_price"] = new_data.get("unit_price") if new_data.get("unit_price") is not None else new_data.get("price")
    new_data["status"] = _compute_status(new_data)

    return ok(data=new_data, message="Medicine updated")


@router.delete("/items/{item_id}", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def delete_item(
    item_id: str,
    current: TokenData = Depends(get_current_active_user),
) -> Any:
    db = get_db()
    ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"]).document(item_id)
    snap = ref.get()
    if not getattr(snap, "exists", False):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Medicine not found")
    now = datetime.utcnow()
    ref.set({"deleted": True, "deleted_at": now, "deleted_by": getattr(current, "user_id", None), "updated_at": now}, merge=True)
    return ok(message="Medicine moved to trash")


@router.get("/trash/items", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def list_trash_items(
    current: TokenData = Depends(get_current_active_user),
    hospital_id: Optional[str] = Query(None),
    q: Optional[str] = Query(None, description="Search by name or SKU"),
    sort: Optional[str] = Query(None, description="recent|oldest|name_asc|name_desc"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
) -> Any:
    db = get_db()
    items_ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"])

    query_ref = items_ref
    if hospital_id:
        query_ref = query_ref.where("hospital_id", "==", hospital_id)

    docs = [d.to_dict() for d in query_ref.stream()]
    docs = [d for d in docs if _is_deleted(d)]

    q_norm = _normalize_str(q).lower()
    if q_norm:
        filtered = []
        for it in docs:
            name = str(it.get("name") or it.get("medicine_name") or "").lower()
            serial = str(it.get("serial_number") or it.get("sku") or "").lower()
            if q_norm in name or q_norm in serial:
                filtered.append(it)
        docs = filtered

    for it in docs:
        it["medicine_name"] = it.get("medicine_name") or it.get("name")
        it["serial_number"] = it.get("serial_number") or it.get("sku")
        it["quantity"] = int(it.get("quantity") if it.get("quantity") is not None else (it.get("stock") or 0))
        it["unit_price"] = it.get("unit_price") if it.get("unit_price") is not None else it.get("price")
        it["status"] = _compute_status(it)

    def _to_dt(v: Any) -> Optional[datetime]:
        try:
            if v is None:
                return None
            if isinstance(v, datetime):
                return v
            to_dt = getattr(v, "to_datetime", None)
            if callable(to_dt):
                return to_dt()
            return datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        except Exception:
            return None

    s = _normalize_str(sort).lower()
    if s == "oldest":
        docs.sort(key=lambda x: _to_dt(x.get("deleted_at") or x.get("updated_at") or x.get("created_at")) or datetime.min)
    elif s == "name_asc":
        docs.sort(key=lambda x: str(x.get("medicine_name") or "").lower())
    elif s == "name_desc":
        docs.sort(key=lambda x: str(x.get("medicine_name") or "").lower(), reverse=True)
    else:
        docs.sort(key=lambda x: _to_dt(x.get("deleted_at") or x.get("updated_at") or x.get("created_at")) or datetime.min, reverse=True)

    total = len(docs)
    start = (page - 1) * page_size
    end = start + page_size
    items = docs[start:end]

    return ok(
        data=items,
        meta={"page": page, "page_size": page_size, "total": total},
    )


@router.post("/trash/items/{item_id}/restore", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def restore_trash_item(
    item_id: str,
    current: TokenData = Depends(get_current_active_user),
) -> Any:
    db = get_db()
    ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"]).document(item_id)
    snap = ref.get()
    if not getattr(snap, "exists", False):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Medicine not found")

    data = snap.to_dict() or {}
    if not _is_deleted(data):
        return ok(message="Medicine is not in trash")

    now = datetime.utcnow()
    ref.set({"deleted": False, "is_deleted": False, "deleted_at": None, "deleted_by": None, "restored_at": now, "restored_by": getattr(current, "user_id", None), "updated_at": now}, merge=True)
    return ok(message="Medicine restored")


@router.delete("/trash/items/{item_id}", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def permanently_delete_trash_item(
    item_id: str,
    current: TokenData = Depends(get_current_active_user),
) -> Any:
    db = get_db()
    ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"]).document(item_id)
    snap = ref.get()
    if not getattr(snap, "exists", False):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Medicine not found")
    ref.delete()
    return ok(message="Medicine permanently deleted")


@router.post("/items/{item_id}/adjust-stock", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def adjust_stock(
    item_id: str,
    payload: Dict[str, Any],
    current: TokenData = Depends(get_current_active_user),
) -> Any:
    try:
        delta = int(payload.get("delta"))
    except Exception:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="delta must be a number")

    reason = _normalize_str(str(payload.get("reason") or ""))
    if not reason:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="reason is required")

    db = get_db()
    item_ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"]).document(item_id)
    snap = item_ref.get()
    if not getattr(snap, "exists", False):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Item not found")

    item = snap.to_dict() or {}
    current_stock = int(item.get("stock") or 0)
    new_stock = current_stock + delta
    if new_stock < 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Not enough stock. Current stock is {current_stock}",
        )

    now = datetime.utcnow()
    item_ref.update({"stock": new_stock, "updated_at": now})

    log_ref = db.collection(COLLECTIONS["PHARMACY_STOCK_LOGS"]).document()
    log_data = {
        "id": log_ref.id,
        "item_id": item_id,
        "hospital_id": item.get("hospital_id"),
        "delta": delta,
        "from_stock": current_stock,
        "to_stock": new_stock,
        "reason": reason,
        "performed_by": current.user_id,
        "created_at": now,
    }
    log_ref.set(log_data)

    return ok(
        data={"item_id": item_id, "stock": new_stock, "log": log_data},
        message="Stock updated",
    )


@router.get("/low-stock", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def low_stock(
    current: TokenData = Depends(get_current_active_user),
    hospital_id: Optional[str] = Query(None),
    threshold: Optional[int] = Query(None, ge=0, description="Override per-item min_stock"),
    limit: int = Query(50, ge=1, le=200),
) -> Any:
    db = get_db()
    items_ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"])

    query_ref = items_ref
    if hospital_id:
        query_ref = query_ref.where("hospital_id", "==", hospital_id)

    docs = [d.to_dict() for d in query_ref.stream()]

    results = []
    for it in docs:
        stock = int(it.get("stock") or 0)
        min_stock = int(it.get("min_stock") or 0)
        t = int(threshold) if threshold is not None else min_stock
        if stock <= t:
            results.append(it)

    results.sort(key=lambda x: int(x.get("stock") or 0))
    results = results[:limit]

    return {"results": results, "count": len(results)}


@public_router.get("/export-excel", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def export_pharmacy_inventory_excel() -> StreamingResponse:
    db = get_db()
    medicines_ref = db.collection(COLLECTIONS.get("PHARMACY_MEDICINES") or "pharmacy_medicines")

    headers = [
        "Medicine Name",
        "Generic Name",
        "Brand",
        "Category",
        "Batch No",
        "Supplier",
        "Purchase Price",
        "Selling Price",
        "Quantity",
        "Unit",
        "Expiry Date",
        "MFG Date",
        "Created At",
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"

    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).font = Font(bold=True)

    def _to_excel_str(v: Any) -> str:
        if v is None:
            return ""
        try:
            if isinstance(v, datetime):
                return v.isoformat()
            to_dt = getattr(v, "to_datetime", None)
            if callable(to_dt):
                dt = to_dt()
                if isinstance(dt, datetime):
                    return dt.isoformat()
        except Exception:
            pass
        try:
            return str(v)
        except Exception:
            return ""

    def _num(v: Any) -> Any:
        if v is None:
            return ""
        return v

    # Paginate for large datasets to avoid loading all docs into memory
    batch_size = 1000
    last_doc = None
    any_rows = False

    try:
        while True:
            q = medicines_ref
            try:
                if hasattr(q, "order_by"):
                    q = q.order_by("__name__")
            except Exception:
                q = medicines_ref

            if last_doc is not None:
                try:
                    if hasattr(q, "start_after"):
                        q = q.start_after(last_doc)
                except Exception:
                    pass

            try:
                q = q.limit(int(batch_size))
            except Exception:
                pass

            docs = list(q.stream())
            if not docs:
                break

            for d in docs:
                m = d.to_dict() or {}
                row = [
                    m.get("name") or "",
                    m.get("generic_name") or "",
                    m.get("brand") or "",
                    m.get("category") or "",
                    m.get("batch_no") or m.get("batch_number") or "",
                    m.get("supplier") or m.get("distributor") or "",
                    _num(m.get("purchase_price")),
                    _num(m.get("selling_price")),
                    _num(m.get("quantity")),
                    m.get("stock_unit") or "",
                    _to_excel_str(m.get("expiration_date")),
                    _to_excel_str(m.get("manufacturing_date")),
                    _to_excel_str(m.get("created_at")),
                ]
                ws.append(row)
                any_rows = True

            last_doc = docs[-1]

            # Safety break if the backend doesn't support pagination correctly
            if len(docs) < batch_size:
                break
    except Exception:
        logger.exception("Failed to export pharmacy inventory")
        raise HTTPException(status_code=status.HTTP_500_INTERNAL_SERVER_ERROR, detail="Failed to export pharmacy inventory")

    # Auto width columns (cap to avoid huge widths)
    try:
        max_lens = [len(h) for h in headers]
        if any_rows:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                for i, val in enumerate(row):
                    try:
                        l = len(str(val)) if val is not None else 0
                    except Exception:
                        l = 0
                    if l > max_lens[i]:
                        max_lens[i] = l
        for i, ml in enumerate(max_lens, start=1):
            ws.column_dimensions[get_column_letter(i)].width = min(max(ml + 2, 12), 50)
    except Exception:
        pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=pharmacy_inventory.xlsx"},
    )


@router.get("/dashboard", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def pharmacy_dashboard(
    current: TokenData = Depends(get_current_active_user),
    hospital_id: Optional[str] = Query(None),
) -> Any:
    db = get_db()
    items_ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"])

    query_ref = items_ref
    if hospital_id:
        query_ref = query_ref.where("hospital_id", "==", hospital_id)

    items = [d.to_dict() for d in query_ref.stream()]

    total_medicines = len(items)
    active_count = 0
    expired_count = 0
    low_stock_count = 0
    inventory_value = 0.0
    total_stock = 0
    med_types: set = set()

    for it in items:
        try:
            qty = int(it.get("quantity") if it.get("quantity") is not None else (it.get("stock") or 0))
        except Exception:
            qty = 0
        try:
            price = float(it.get("unit_price") if it.get("unit_price") is not None else (it.get("price") or 0))
        except Exception:
            price = 0.0
        try:
            min_stock = int(it.get("min_stock") or 0)
        except Exception:
            min_stock = 0

        st = _compute_status(it)
        if st == "expired":
            expired_count += 1
        elif st == "active":
            active_count += 1

        if qty <= min_stock:
            low_stock_count += 1

        total_stock += max(qty, 0)
        tname = str(it.get("type") or it.get("category") or "").strip()
        if tname:
            med_types.add(tname.lower())

        inventory_value += max(qty, 0) * max(price, 0.0)

    alerts = []
    if low_stock_count > 0:
        alerts.append({
            "type": "low_stock",
            "title": "Low Stock Alert",
            "message": f"{low_stock_count} item{'s' if low_stock_count != 1 else ''} has low stock. Please reorder soon.",
            "count": low_stock_count,
        })
    if expired_count > 0:
        alerts.append({
            "type": "expired",
            "title": "Expired Items",
            "message": f"{expired_count} item{'s' if expired_count != 1 else ''} have expired. Please remove from inventory.",
            "count": expired_count,
        })

    return ok(
        data={
            "total_medicines": total_medicines,
            "active_medicines": active_count,
            "low_stock_items": low_stock_count,
            "expired_items": expired_count,
            "total_stock": int(total_stock),
            "medicine_types": len(med_types),
            "inventory_value": round(inventory_value, 2),
            "currency": "PKR",
            "alerts": alerts,
        }
    )


@router.post("/sales", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def record_sale(
    payload: Dict[str, Any],
    current: TokenData = Depends(get_current_active_user),
) -> Any:
    """Record a pharmacy sale transaction.

    Expected fields from frontend:
    - hospital_id (required)
    - item_id (optional but recommended)
    - medicine_name (optional when item_id provided)
    - quantity_sold (required)
    - unit_price (optional, defaults to item.unit_price)
    - sold_at (optional date string)
    """
    hospital_id = _normalize_str(str(payload.get("hospital_id") or ""))
    if not hospital_id:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="hospital_id is required")

    item_id = _normalize_str(str(payload.get("item_id") or "")) or None
    quantity_sold = _coerce_int(payload.get("quantity_sold"), field="quantity_sold")
    if quantity_sold <= 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="quantity_sold must be greater than 0")

    sold_at_iso = None
    if payload.get("sold_at"):
        sold_at_iso = _normalize_date_str(str(payload.get("sold_at")))

    payment_status = _normalize_str(str(payload.get("payment_status") or "paid")).lower()
    if payment_status not in {"paid", "pending", "failed", "refunded"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="payment_status is invalid")

    db = get_db()

    medicine_name = _normalize_str(str(payload.get("medicine_name") or ""))
    unit_price_val: Optional[float] = None
    if payload.get("unit_price") is not None:
        unit_price_val = _coerce_float(payload.get("unit_price"), field="unit_price")

    # If item_id is provided, use inventory to fill details + decrement stock
    if item_id:
        item_ref = db.collection(COLLECTIONS["PHARMACY_ITEMS"]).document(item_id)
        snap = item_ref.get()
        if not getattr(snap, "exists", False):
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Medicine not found in inventory")
        item = snap.to_dict() or {}

        if hospital_id and str(item.get("hospital_id") or "") != hospital_id:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Medicine does not belong to this hospital")

        medicine_name = medicine_name or str(item.get("medicine_name") or item.get("name") or "")
        if unit_price_val is None:
            try:
                unit_price_val = float(item.get("unit_price") if item.get("unit_price") is not None else (item.get("price") or 0))
            except Exception:
                unit_price_val = 0.0

        current_qty = int(item.get("quantity") if item.get("quantity") is not None else (item.get("stock") or 0))
        new_qty = current_qty - quantity_sold
        if new_qty < 0:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Not enough stock. Current stock is {current_qty}")

        now = datetime.utcnow()
        item_ref.update({"quantity": new_qty, "stock": new_qty, "updated_at": now})

        # Stock log for sale
        log_ref = db.collection(COLLECTIONS["PHARMACY_STOCK_LOGS"]).document()
        log_ref.set({
            "id": log_ref.id,
            "item_id": item_id,
            "hospital_id": hospital_id,
            "delta": -quantity_sold,
            "from_stock": current_qty,
            "to_stock": new_qty,
            "reason": _normalize_str(str(payload.get("reason") or "sale")),
            "performed_by": current.user_id,
            "created_at": now,
        })
    else:
        if not medicine_name:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="medicine_name is required when item_id is not provided")
        if unit_price_val is None:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="unit_price is required when item_id is not provided")

    unit_price_val = float(unit_price_val or 0.0)
    if unit_price_val < 0:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="unit_price cannot be negative")

    total_amount = round(quantity_sold * unit_price_val, 2)
    now = datetime.utcnow()
    dispensed_at = now
    if sold_at_iso:
        try:
            dispensed_at = datetime.fromisoformat(sold_at_iso)
        except Exception:
            dispensed_at = now

    sales_ref = db.collection(COLLECTIONS["PHARMACY_SALES"])
    sale_doc = sales_ref.document()
    sale_data = {
        "id": sale_doc.id,
        "hospital_id": hospital_id,
        "item_id": item_id,
        "medicine_name": medicine_name,
        "quantity_sold": quantity_sold,
        "unit_price": unit_price_val,
        "total_amount": total_amount,
        # Keep the UI-friendly date field
        "sold_at": sold_at_iso or now.date().isoformat(),
        # Accounting fields
        "payment_status": payment_status,
        "dispensed_at": dispensed_at,
        "created_at": now,
        "performed_by": current.user_id,
    }
    sale_doc.set(sale_data)

    return ok(data=sale_data, message="Sale recorded")


@router.get("/sales", dependencies=[Depends(require_roles("pharmacy", "admin"))])
async def sales_and_revenue(
    current: TokenData = Depends(get_current_active_user),
    hospital_id: Optional[str] = Query(None),
    period: str = Query("this_week", description="today|this_week|this_month|custom|all"),
    start_date: Optional[str] = Query(None, description="For custom range (DD/MM/YYYY or YYYY-MM-DD)"),
    end_date: Optional[str] = Query(None, description="For custom range (DD/MM/YYYY or YYYY-MM-DD)"),
    payment_status: str = Query("paid", description="paid|pending|failed|refunded|all"),
    q: Optional[str] = Query(None, description="Search by medicine name"),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
) -> Any:
    """Sales & Revenue page API.

    Returns:
    - weekly_sales (last 7 days)
    - monthly_sales (this month)
    - total_revenue (all time)
    - sales_history (filtered by period/custom range)
    """
    db = get_db()
    sales_ref = db.collection(COLLECTIONS["PHARMACY_SALES"])

    status_norm = _normalize_str(payment_status).lower()
    if status_norm not in {"paid", "pending", "failed", "refunded", "all"}:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="payment_status is invalid")

    def _base_query():
        qref = sales_ref
        if hospital_id:
            qref = qref.where("hospital_id", "==", hospital_id)
        if status_norm != "all":
            qref = qref.where("payment_status", "==", status_norm)
        return qref

    def _sum_query(qref) -> float:
        total = 0.0
        for d in qref.stream():
            data = d.to_dict() or {}
            try:
                total += float(data.get("total_amount") or 0)
            except Exception:
                continue
        return round(total, 2)

    now = datetime.utcnow()
    today_start = datetime(now.year, now.month, now.day)
    week_start = now - timedelta(days=7)
    month_start = datetime(now.year, now.month, 1)

    total_revenue = _sum_query(_base_query())
    today_sales = _sum_query(_base_query().where("dispensed_at", ">=", today_start))
    weekly_sales = _sum_query(_base_query().where("dispensed_at", ">=", week_start))
    monthly_sales = _sum_query(_base_query().where("dispensed_at", ">=", month_start))

    prev_week_sales = _sum_query(
        _base_query().where("dispensed_at", ">=", week_start - timedelta(days=7)).where("dispensed_at", "<=", _end_inclusive(week_start))
    )

    prev_month_end = month_start
    prev_month_start = datetime((month_start - timedelta(days=1)).year, (month_start - timedelta(days=1)).month, 1)
    prev_month_sales = _sum_query(
        _base_query().where("dispensed_at", ">=", prev_month_start).where("dispensed_at", "<=", _end_inclusive(prev_month_end))
    )

    def _pct_change(current_val: float, prev_val: float) -> Optional[float]:
        if prev_val <= 0:
            return None
        return round(((current_val - prev_val) / prev_val) * 100.0, 2)

    weekly_change_pct = _pct_change(weekly_sales, prev_week_sales)
    monthly_change_pct = _pct_change(monthly_sales, prev_month_sales)

    start, end = _period_range(period, start_date, end_date)
    hist_query = _base_query()
    if start:
        hist_query = hist_query.where("dispensed_at", ">=", start)
    if end:
        hist_query = hist_query.where("dispensed_at", "<=", _end_inclusive(end))

    docs = [d.to_dict() for d in hist_query.stream()]

    items: List[Dict[str, Any]] = []
    for s in docs:
        dispensed_dt = _to_datetime(s.get("dispensed_at"))
        if dispensed_dt is None:
            dispensed_dt = _to_datetime(s.get("created_at"))
        s["_sold_dt"] = dispensed_dt
        try:
            s["total_amount"] = float(s.get("total_amount") or 0)
        except Exception:
            s["total_amount"] = 0.0
        try:
            s["unit_price"] = float(s.get("unit_price") or 0)
        except Exception:
            s["unit_price"] = 0.0
        try:
            s["quantity_sold"] = int(s.get("quantity_sold") or 0)
        except Exception:
            s["quantity_sold"] = 0
        s["date"] = (dispensed_dt.date().isoformat() if dispensed_dt else (str(s.get("sold_at") or "") or None))
        items.append(s)

    q_norm = _normalize_str(q).lower()
    if q_norm:
        items = [it for it in items if q_norm in str(it.get("medicine_name") or "").lower()]

    items.sort(key=lambda x: x.get("_sold_dt") or datetime.min, reverse=True)

    total_filtered = len(items)
    start_i = (page - 1) * page_size
    end_i = start_i + page_size
    page_items = items[start_i:end_i]

    history = []
    for it in page_items:
        history.append({
            "id": it.get("id"),
            "date": it.get("date"),
            "medicine_name": it.get("medicine_name"),
            "quantity_sold": it.get("quantity_sold"),
            "unit_price": it.get("unit_price"),
            "total_amount": it.get("total_amount"),
            "payment_status": it.get("payment_status"),
        })

    return ok(
        data={
            "today_sales": today_sales,
            "weekly_sales": weekly_sales,
            "weekly_change_pct": weekly_change_pct,
            "monthly_sales": monthly_sales,
            "monthly_change_pct": monthly_change_pct,
            "total_revenue": total_revenue,
            "currency": "PKR",
            "period": _normalize_str(period) or "this_week",
            "payment_status": status_norm,
            "sales_history": history,
        },
        meta={"page": page, "page_size": page_size, "total": total_filtered},
    )
